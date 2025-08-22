from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import shutil
import os

# Путь до файла-источника и до конечного файла
Path = (os.path.abspath(__file__)) # путь к питон файлу
SourcePath = str(os.path.dirname(os.path.dirname(Path))) + "/buffer/buffer.xlsm" # путь к buffer
TargetPath = str(os.path.dirname(os.path.dirname(os.path.dirname(Path)))) + "/result.xlsm" # путь к result
CleanedPath = str(os.path.dirname(Path)) + "/cleaned.xlsm"


# Объявляем допустимые столбцы
SetColumns = {("ISIN", "0"):1, ("Ticker &", "Exchange"):2, ("Ccy", "0"):3, ("Cpn", "(%)"):4, ("0", "(%)"):4, ("Name", "1"):5, ("Sector", "0"):6, ("Industry", "0"):7, ("Maturity", "(1. call date)"):8, ("Price", "MID"):9, ("Price", "1"):9, ("Mk-Cap", "mia"):10, ("YTM", "MID"):11, ("Share class", "0"):12, ("Share", "class"):12, ("ER/MF", "0"):13, ("Rating", "Moody"):14, ("Rating", "S&P"):15, ("Rating", "Fitch"):16, ("Size", "mio"):17, ("Z-", "spread"):18, ("ASW", "spread"):19, ("Min", "piece"):20, ("Min", "incr"):21, (0, "Mkt of Issue"):22, ("Notes", "0"):23, (0, "Notes"):23, ("Added on", "0"):24, ("Added", "on"):24}
# Объявляем стобцы в конечном файле (только для конечного файла)
Columns = (("ISIN", 0), ("Ticker &", "Exchange"), ("Ccy", 0), ("Cpn", "(%)"), ("Name", 0), ("Sector", 0), ("Industry", 0), ("Maturity", "(1. call date)"), ("Price", 0), ("Mk-Cap", "mia"), ("YTM", "MID"), ("Share", "classes"), ("ER/MF", 0), ("Rating", "Mood"), ("Rating", "S&P"), ("Rating", "Fitch"), ("Size", "mio"), ("Z-", "Spread"), ("ASW", "spread"), ("Min", "piece"), ("Min", "incr"), ("Mkt of", "Issue"), ("Notes", 0), ("Added", "on"))
Used_positions = set()

def data_finding(SourcePath):
    TARGET_COLOR = "FFC0E9C0"
    base_wb = load_workbook(SourcePath, read_only=True)
    Chosen_assets = {"Без названия": set()}

    # Здесь идём по листам - далее всё для конкретного листа
    for base_ws in base_wb.worksheets:
        current_title = "Без названия"
        downer_row = base_ws.max_row
        leftern_column = base_ws.max_column
        dic_to_copy = {}
        
        # FIXED: Extract values from cells instead of using cell objects
        for title_column in range(3, leftern_column+1):
            cell1 = base_ws.cell(row=2, column=title_column)
            cell2 = base_ws.cell(row=3, column=title_column)
            # Get values from cells, handle None values
            value1 = cell1.value if cell1.value is not None else "0"
            value2 = cell2.value if cell2.value is not None else "0"
            title_to_check = (value1, value2)
            title_to_check2 = (value1, "1")
            
            if (title_to_check in SetColumns):
                dic_to_copy[title_column] = SetColumns[title_to_check]
            if (title_to_check2 in SetColumns):
                dic_to_copy[title_column] = SetColumns[title_to_check2]

        # Ищем нужные строки (идём по столбцу C)
        for row in base_ws.iter_rows(min_row=5, max_row=downer_row, min_col=3, max_col=3):
            cell_C = row[0]
            
            # FIXED: Handle EmptyCell objects
            if hasattr(cell_C, 'row'):
                Curr_row = cell_C.row
            else:
                # If it's an EmptyCell, skip this row
                continue
                
            value_C = cell_C.value
            
            # Проверка цвета на зелёный
            has_special_color = False
            if (hasattr(cell_C, 'fill') and 
                cell_C.fill and 
                cell_C.fill.fgColor and 
                cell_C.fill.fgColor.rgb):
                has_special_color = (cell_C.fill.fgColor.rgb == TARGET_COLOR)

            # Если пусто - скипни
            if not value_C:
                continue

            # Если упали сюда - значение не пусто
            str_value_C = str(value_C)
            
            if len(str_value_C) != 12: # значит это название группы
                if str_value_C not in Chosen_assets:
                    Chosen_assets[str_value_C] = set()
                current_title = str_value_C
            else: # значит это бумага
                if has_special_color: # выделенная бумага, нужна нам
                    # Read entire row values at once
                    row_values = next(base_ws.iter_rows(min_row=Curr_row, max_row=Curr_row, 
                                                    values_only=True))
                    
                    # Create list with 24 None values
                    list_of_data = [None] * 24
                    
                    # Fill only the positions specified in dic_to_copy
                    for source_col, target_pos in dic_to_copy.items():
                        if source_col <= len(row_values):
                            list_of_data[target_pos-1] = row_values[source_col-1]
                            Used_positions.add(target_pos)
                    Chosen_assets[current_title].add(tuple(list_of_data))
    
    base_wb.close()
    return Chosen_assets

def insert_dict_to_excel_fast(dictionary, excel_file_path):
    # Загружаем существующий Excel файл
    wb = load_workbook(excel_file_path)
    ws = wb.active
    LastColumn = len(Used_positions)
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column = LastColumn)
    ws.cell(row=1, column=2, value="Balanced Portfolio")
    ws.cell(row=1, column=2).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=1, column=2).font = Font(size=40, bold=True, color='808080', name='Calabria Light')
    current_row = 4
    
    Counter_for_titles = 1
    for Num_for_main_titles in range(len(Columns)):
        if Num_for_main_titles+1 not in Empty_columns:
            ws.cell(row=2, column=Counter_for_titles).value = Columns[Num_for_main_titles][0] if Columns[Num_for_main_titles][0] != 0 else None
            ws.cell(row=3, column=Counter_for_titles).value = Columns[Num_for_main_titles][1] if Columns[Num_for_main_titles][1] != 0 else None
            Counter_for_titles += 1

    FONT = Font(bold=True, color='808080', name='Calabria Light')
    FILL = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
    BORDER = Border(left=Side(style=None), right=Side(style=None), top=Side(style='thin'), bottom=Side(style='thin'))
    LEFT = Alignment(horizontal='left', vertical='bottom')
    
    for key, value_set in dictionary.items():
        # Пропускаем пустые множества
        if not value_set:
            continue
            
        # Вставляем ключ как отдельную строку
        ws.cell(row=current_row, column=1, value=key)
        for i in range(1, LastColumn+1):
            ws.cell(row=current_row, column=i).font = FONT
            ws.cell(row=current_row, column=i).fill = FILL
        current_row += 1
        
        # Вставляем данные из кортежей
        for tuple_item in value_set:
            i = 0
            for col_idx, value in enumerate(tuple_item, 1):
                if col_idx in Empty_columns:
                    i += 1
                    continue
                # Вставляем значение, если не None
                if value is not None:
                    ws.cell(row=current_row, column=col_idx-i, value=value)
                    ws.cell(row=current_row, column=col_idx-i).alignment = LEFT
                    if isinstance (value, float):
                        ws.cell(row=current_row, column=col_idx-i).number_format = '#,##0.0'
            current_row += 1

        for i in range(1, LastColumn+1):
            Column_letter = get_column_letter(i)
            ws.column_dimensions[Column_letter].auto_size = True
            for j in range(4, ws.max_row+1):
                ws.cell(row = j, column = i).border = BORDER

    # Сохраняем файл
    wb.save(excel_file_path)
    wb.close

# Использование

dictionary = data_finding(SourcePath)
shutil.copy2(CleanedPath, TargetPath)
Empty_columns = set(range(1, 25)) - Used_positions
insert_dict_to_excel_fast(dictionary, TargetPath)
